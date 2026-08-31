#!/usr/bin/env python3
"""Export an xlsx workbook to a single, self-contained HTML file.

The HTML opens in any browser (no Excel needed) and preserves the per-class
colors, the grading block, and the clickable links to the detail .md files.

Usage:
    python export_html.py <workbook.xlsx> [--out <output.html>]
"""
from __future__ import annotations

import argparse
import html
from pathlib import Path

from openpyxl import load_workbook


def argb_to_css(color, default="#ffffff"):
    if color is None:
        return default
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return default
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return default
    return f"#{rgb.lower()}"


def text_color(hex_color: str) -> str:
    try:
        r = int(hex_color[1:3], 16)
        g = int(hex_color[3:5], 16)
        b = int(hex_color[5:7], 16)
    except (ValueError, IndexError):
        return "#1e1e1e"
    lum = 0.299 * r + 0.587 * g + 0.114 * b
    return "#ffffff" if lum < 140 else "#1e1e1e"


def render_sheet(ws) -> str:
    covered = set()
    span = {}
    for mr in ws.merged_cells.ranges:
        span[(mr.min_row, mr.min_col)] = (
            mr.max_col - mr.min_col + 1,
            mr.max_row - mr.min_row + 1,
        )
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    covered.add((r, c))

    tab = argb_to_css(ws.sheet_properties.tabColor, "#cccccc")
    out = [f'<section><h2 style="border-left:14px solid {tab}">{html.escape(ws.title)}</h2>']
    out.append("<table>")
    for r in range(1, ws.max_row + 1):
        out.append("<tr>")
        for c in range(1, ws.max_column + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            bg = "#ffffff"
            if cell.fill and cell.fill.patternType == "solid":
                bg = argb_to_css(cell.fill.fgColor, "#ffffff")
            fg = text_color(bg)
            styles = [f"background:{bg}", f"color:{fg}"]
            if cell.font and cell.font.bold:
                styles.append("font-weight:700")
            colspan, rowspan = span.get((r, c), (1, 1))
            attrs = f' colspan="{colspan}"' if colspan > 1 else ""
            attrs += f' rowspan="{rowspan}"' if rowspan > 1 else ""
            val = "" if cell.value is None else str(cell.value)
            if cell.hyperlink and cell.hyperlink.target:
                target = html.escape(cell.hyperlink.target, quote=True)
                inner = f'<a href="{target}">{html.escape(val)}</a>'
            else:
                inner = html.escape(val)
            out.append(f'<td{attrs} style="{";".join(styles)}">{inner}</td>')
        out.append("</tr>")
    out.append("</table></section>")
    return "\n".join(out)


def export(workbook_path: Path, out_path: Path) -> None:
    wb = load_workbook(str(workbook_path))
    body = "\n".join(render_sheet(wb[name]) for name in wb.sheetnames)
    doc = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(workbook_path.stem)}</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif; margin: 24px; color:#1e1e1e; }}
  section {{ margin-bottom: 40px; }}
  h2 {{ padding-left: 10px; }}
  table {{ border-collapse: collapse; width: 100%; max-width: 1100px; }}
  td {{ border: 1px solid #d9d9d9; padding: 6px 10px; font-size: 14px; vertical-align: top; }}
  a {{ color: #0563C1; }}
</style>
</head>
<body>
<h1>{html.escape(workbook_path.stem)}</h1>
{body}
</body>
</html>
"""
    out_path.write_text(doc, encoding="utf-8")
    print(f"Wrote {out_path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export xlsx to standalone HTML.")
    parser.add_argument("workbook", help="Path to the .xlsx workbook")
    parser.add_argument("--out", help="Output .html path (defaults next to workbook)")
    args = parser.parse_args()

    wb_path = Path(args.workbook).expanduser()
    out_path = Path(args.out).expanduser() if args.out else wb_path.with_suffix(".html")
    export(wb_path, out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
