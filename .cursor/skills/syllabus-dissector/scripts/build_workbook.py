#!/usr/bin/env python3
"""Append a dissected syllabus to a color-coded Excel workbook.

Reads a JSON description of one class (see schema below), then:
  * creates/updates a per-class worksheet (data is "separated by class"),
  * writes the grading scale and the exact threshold required for an 'A',
  * lists each grading category sorted by weight (highest first),
  * records start date, due date, and group-project flag per category,
  * generates a detailed PDF per assignment (all syllabus info extracted),
  * hyperlinks each row to its PDF document,
  * color-codes the sheet tab and rows based on the class.

Re-running for the same class replaces that class's sheet and PDFs (idempotent).

Usage:
    python build_workbook.py <class.json> --workbook syllabi.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Import PDF generator from sibling script
sys.path.insert(0, str(Path(__file__).resolve().parent))
from generate_pdfs import generate_all, slugify  # noqa: E402

PALETTE = [
    "1F77B4", "D62728", "2CA02C", "9467BD",
    "FF7F0E", "17BECF", "8C564B", "E377C2",
]

HEADERS = [
    "Category", "Weight", "Start Date", "Due Date", "Group Project", "Details",
]

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

OVERVIEW_SHEET = "Overview"
CONNECT_PREP_MARK = "Personal Assessments (Connect)"
OVERVIEW_AS_OF: date | None = None
TODO_LABELS = frozenset({"Homework", "Assignment", "Exam", "Certification", "Research"})
MAJOR_LABELS = frozenset({"Assignment", "Exam", "Certification"})
MAJOR_NAME = re.compile(
    r"\b(paper|midterm|final exam|final|exam|project|presentation|memo|"
    r"certification|simternship|linkedin|proposal|contract|reflection)\b",
    re.I,
)
PAPER_NAME = re.compile(r"\bpaper\b", re.I)
EXTERNAL_CERT_NAME = re.compile(
    r"\b(linkedin|linked in|stukent|simternship|certification)\b",
    re.I,
)
MONTH_NAMES = (
    "",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
)

# Sub-row type labels (Category column) — sort order for grouped display
SUB_LABEL_ORDER = {
    "Reading": 0,
    "Homework": 1,
    "Assignment": 2,
    "Exam": 3,
    "Research": 4,
    "Certification": 5,
}
SUB_LABEL_COLORS = {
    "Reading": "1F4E79",       # navy
    "Homework": "C65911",      # orange
    "Assignment": "7030A0",    # purple
    "Exam": "C00000",          # red
    "Research": "375623",      # green
    "Certification": "0070C0", # blue
}


def clean_hex(value: str) -> str:
    value = (value or "").lstrip("#").strip().upper()
    if len(value) == 8:
        value = value[2:]
    if not re.fullmatch(r"[0-9A-F]{6}", value):
        raise ValueError(f"Invalid hex color: {value!r}")
    return value


def auto_color(class_code: str) -> str:
    return PALETTE[sum(ord(c) for c in class_code) % len(PALETTE)]


def tint(hex_color: str, factor: float) -> str:
    r, g, b = (int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]", "-", name).strip() or "Sheet"
    return name[:31]


def sort_sub_assignments(items: list[dict]) -> list[dict]:
    """Chronological order (earliest first); Reading immediately before Homework on same due date."""
    def session_date(item: dict) -> str:
        return item.get("due_date") or item.get("start_date") or ""

    def key(item: dict) -> tuple:
        label = item.get("notes") or ""
        return (session_date(item), SUB_LABEL_ORDER.get(label, 99), item.get("name") or "")

    return sorted(items, key=key)


def weight_breakdown_text(categories: list[dict]) -> str:
    """One-line grade weight summary for the sheet header bar."""
    parts: list[str] = []
    for cat in sorted(categories, key=weight_sort_key):
        w = cat.get("weight")
        if w is None:
            continue
        parts.append(f"{cat.get('name', '')} {format_weight(cat)}")
    return "  |  ".join(parts)


def format_weight(cat: dict) -> str:
    w = cat.get("weight")
    if w is None:
        return "—"
    unit = (cat.get("weight_unit") or "percent").lower()
    if unit.startswith("point") or unit in {"pts", "pt"}:
        return f"{w:g} pts"
    return f"{w:g}%"


def weight_sort_key(cat: dict):
    w = cat.get("weight")
    if w is None:
        return (1, 0.0)
    try:
        return (0, -float(w))
    except (TypeError, ValueError):
        return (1, 0.0)


def remove_class_sheet(wb: Workbook, class_code: str) -> None:
    name = sanitize_sheet_name(class_code)
    if name in wb.sheetnames:
        wb.remove(wb[name])


def remove_old_detail_sheets(wb: Workbook) -> None:
    """Remove leftover detail tabs from an older workbook format."""
    for name in list(wb.sheetnames):
        if "-" in name and name not in {sanitize_sheet_name(n) for n in wb.sheetnames}:
            # detail sheets look like "HIST-103-Sleep Paper"
            parts = name.split("-", 1)
            if len(parts) == 2 and any(
                name.startswith(f"{sanitize_sheet_name(c)}-")
                for c in wb.sheetnames
                if "-" not in c or c == sanitize_sheet_name(c)
            ):
                pass  # handled below
    # Remove any sheet whose name starts with "<ClassCode>-" except we keep summary sheets
    summary = {n for n in wb.sheetnames if "-" not in n or n.count("-") == 0}
    # Summary sheets are like HIST-103 (one hyphen) - detail sheets have more content after
    to_remove = []
    for name in wb.sheetnames:
        for summary_name in wb.sheetnames:
            if name != summary_name and name.startswith(f"{summary_name}-"):
                to_remove.append(name)
    for name in to_remove:
        if name in wb.sheetnames:
            wb.remove(wb[name])


def remove_class_pdfs(docs_dir: Path, class_code: str) -> None:
    prefix = slugify(class_code) + "-"
    if docs_dir.exists():
        for f in docs_dir.glob("*.pdf"):
            if f.stem.startswith(prefix):
                f.unlink()


def remove_blank_template_sheet(wb: Workbook) -> None:
    if "Sheet" not in wb.sheetnames or len(wb.sheetnames) <= 1:
        return
    ws = wb["Sheet"]
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        wb.remove(ws)


def default_template() -> Path:
    return Path(__file__).resolve().parent.parent / "templates" / "blank_template.xlsx"


def pdf_hyperlink(target: str, label: str = "Open PDF") -> str:
    """Excel formula that opens a URL or file path when clicked."""
    safe = target.replace('"', '""')
    return f'=HYPERLINK("{safe}","{label}")'


def sheet_link(sheet_name: str, label: str | None = None) -> str:
    """Internal workbook hyperlink to another worksheet."""
    safe_sheet = sheet_name.replace("'", "''")
    text = label or sheet_name
    safe_label = text.replace('"', '""')
    return f'=HYPERLINK("#\'{safe_sheet}\'!A1","{safe_label}")'


def collect_class_json_files(workbook_path: Path) -> list[dict]:
    """Load every class JSON next to the workbook (output/*.json)."""
    json_dir = workbook_path.parent
    out: list[dict] = []
    for path in sorted(json_dir.glob("*.json")):
        if path.name == "rmp.json":
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        if data.get("class", {}).get("code") and data.get("categories") is not None:
            out.append(data)
    return out


def load_rmp_ratings(workbook_path: Path) -> dict[str, dict]:
    """Load Rate My Professors data from output/rmp.json (keyed by class code)."""
    rmp_path = workbook_path.parent / "rmp.json"
    if not rmp_path.exists():
        return {}
    try:
        raw = json.loads(rmp_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    return raw if isinstance(raw, dict) else {}


def rmp_for_class(
    class_code: str,
    instructor: str,
    rmp_table: dict[str, dict],
    class_rmp: dict | None = None,
) -> dict | None:
    """Merge class-level rmp block with shared rmp.json lookup."""
    entry = dict(rmp_table.get(class_code) or {})
    if class_rmp:
        entry.update(class_rmp)
    if not entry:
        return None
    if instructor and not entry.get("instructor"):
        entry["instructor"] = instructor
    return entry if entry.get("quality") is not None else None


def format_rmp_instructor_label(rmp: dict, fallback: str = "") -> str:
    """Syllabus instructor(s) plus RMP profile name when they differ."""
    syllabus = (rmp.get("instructor") or fallback or "").strip()
    profile = (rmp.get("rmp_instructor") or "").strip()
    if profile and profile != syllabus and syllabus:
        return f"{syllabus} (RMP: {profile})"
    return profile or syllabus or "-"


def format_avg_reviewer_grade(rmp: dict) -> str:
    """Self-reported letter-grade average from RMP reviewers."""
    letter = rmp.get("avg_grade_letter")
    gpa = rmp.get("avg_grade_gpa")
    reported = rmp.get("grades_reported")
    total = rmp.get("num_ratings")
    if not letter or gpa is None:
        return "-"
    text = f"{letter} ({gpa:g})"
    if reported and total:
        text += f" [{reported}/{total} reported]"
    elif reported:
        text += f" [{reported} reported]"
    return text


def _due_sort_key(due: str, label: str, name: str, is_sub: bool = False) -> tuple:
    return (due or "9999-99-99", 1 if is_sub else 0, SUB_LABEL_ORDER.get(label, 99), name)


def is_todo_item(item: dict) -> bool:
    """Assignments to do - excludes Reading prep."""
    if item.get("label") == "Reading":
        return False
    if item.get("label") in TODO_LABELS:
        return True
    if not item.get("is_sub") and item.get("is_major"):
        return True
    return False


def _todo_sort_key(item: dict) -> tuple:
    """Prefer labeled assignment sub-rows over parent category rows on the same day."""
    label = item.get("label") or ""
    return (
        item.get("due_date") or "9999-99-99",
        0 if label in TODO_LABELS else 1,
        1 if item.get("is_sub") else 0,
        SUB_LABEL_ORDER.get(label, 99),
        item.get("name") or "",
    )


def first_todo_item(due_items: list[dict]) -> dict | None:
    todos = [x for x in due_items if is_todo_item(x)]
    if not todos:
        return None
    todos.sort(key=_todo_sort_key)
    return todos[0]


def first_todo_by_class(all_due: list[dict]) -> list[dict]:
    """Earliest assignment (non-reading) per class."""
    todos = [x for x in all_due if is_todo_item(x)]
    todos.sort(key=_todo_sort_key)
    seen: set[str] = set()
    out: list[dict] = []
    for item in todos:
        code = item["class_code"]
        if code in seen:
            continue
        seen.add(code)
        out.append(item)
    return out


def build_do_first_items(all_due: list[dict], focus_month: str) -> tuple[list[dict], list[dict]]:
    """Graded assignments + Connect prep: first per class always included, plus focus-month items."""
    todos = [x for x in all_due if is_todo_item(x)]
    graded = [x for x in todos if not x.get("is_connect_prep")]
    connect = [x for x in todos if x.get("is_connect_prep")]
    month_graded = [x for x in graded if _month_key(x["due_date"]) == focus_month]
    month_connect = [x for x in connect if _month_key(x["due_date"]) == focus_month]

    def merge(first_pool: list[dict], month_pool: list[dict]) -> list[dict]:
        merged: list[dict] = []
        seen: set[tuple[str, str, str]] = set()

        def add(item: dict) -> None:
            key = (item["class_code"], item["due_date"], item["name"])
            if key not in seen:
                seen.add(key)
                merged.append(item)

        for item in first_pool:
            add(item)
        for item in month_pool:
            add(item)
        merged.sort(key=_todo_sort_key)
        return merged

    first_graded = [x for x in first_todo_by_class(all_due) if not x.get("is_connect_prep")]
    first_connect = [x for x in first_todo_by_class(all_due) if x.get("is_connect_prep")]
    return merge(first_graded, month_graded), merge(first_connect, month_connect)


def upcoming_todo_items(all_due: list[dict], after_month: str, days: int = 45, *, as_of: date | None = None) -> list[dict]:
    """Graded assignments due after focus month within the next N days."""
    todos = [x for x in all_due if is_todo_item(x) and x.get("is_graded", True) and not x.get("is_connect_prep")]
    if not todos:
        return []
    start = as_of or date.today()
    end = start + timedelta(days=days)
    out: list[dict] = []
    for item in todos:
        iso = item.get("due_date") or ""
        if not iso or _month_key(iso) == after_month:
            continue
        try:
            d = date.fromisoformat(iso)
        except ValueError:
            continue
        if start <= d <= end:
            out.append(item)
    out.sort(key=_todo_sort_key)
    return out


def _is_major_category(cat: dict) -> bool:
    name = cat.get("name") or ""
    if cat.get("weight") is None and not MAJOR_NAME.search(name):
        return False
    return bool(MAJOR_NAME.search(name))


def _paper_names(categories: list[dict]) -> list[str]:
    return [
        cat.get("name") or ""
        for cat in categories
        if PAPER_NAME.search(cat.get("name") or "")
    ]


def _external_certifications(categories: list[dict]) -> list[str]:
    """Graded third-party certs (LinkedIn Learning, Stukent, etc.) - not Connect prep."""
    names: list[str] = []
    for cat in categories:
        if cat.get("weight") is None:
            continue
        cname = cat.get("name") or ""
        has_cert_sub = any(
            sub.get("notes") == "Certification" for sub in cat.get("assignments") or []
        )
        if has_cert_sub or EXTERNAL_CERT_NAME.search(cname):
            names.append(cname)
    return names


def iter_due_items(data: dict) -> list[dict]:
    """Flatten category + sub-rows that carry due dates."""
    code = str(data.get("class", {}).get("code") or "")
    items: list[dict] = []
    for cat in data.get("categories") or []:
        cat_name = cat.get("name") or ""
        cat_weight = cat.get("weight")
        is_connect_prep = cat_name == CONNECT_PREP_MARK or (
            "connect" in cat_name.lower() and cat_weight is None
        )
        is_graded_cat = cat_weight is not None and not is_connect_prep
        due = cat.get("due_date") or ""
        if due:
            items.append(
                {
                    "class_code": code,
                    "category": cat_name,
                    "name": cat_name,
                    "due_date": due,
                    "label": "",
                    "is_major": _is_major_category(cat),
                    "is_sub": False,
                    "is_graded": is_graded_cat,
                    "is_connect_prep": is_connect_prep,
                }
            )
        for sub in cat.get("assignments") or []:
            sd = sub.get("due_date") or ""
            if not sd:
                continue
            label = sub.get("notes") or ""
            sname = sub.get("name") or ""
            is_major = label in MAJOR_LABELS or (
                label == "Assignment" and bool(MAJOR_NAME.search(sname))
            )
            items.append(
                {
                    "class_code": code,
                    "category": cat_name,
                    "name": sname,
                    "due_date": sd,
                    "label": label,
                    "is_major": is_major,
                    "is_sub": True,
                    "is_graded": is_graded_cat,
                    "is_connect_prep": is_connect_prep,
                }
            )
    return items


def class_has_label(data: dict, label: str) -> bool:
    for cat in data.get("categories") or []:
        for sub in cat.get("assignments") or []:
            if sub.get("notes") == label:
                return True
    return False


def class_has_connect_prep(categories: list[dict]) -> bool:
    return any(
        (c.get("name") or "") == CONNECT_PREP_MARK
        or ("connect" in (c.get("name") or "").lower() and c.get("weight") is None)
        for c in categories
    )


def summarize_class(data: dict) -> dict:
    cls = data.get("class", {})
    code = str(cls.get("code") or "")
    categories = data.get("categories") or []
    due_items = iter_due_items(data)
    sorted_major = sorted(
        [x for x in due_items if x["is_major"]],
        key=lambda x: _due_sort_key(x["due_date"], x["label"], x["name"], x["is_sub"]),
    )
    first_todo = first_todo_item(due_items)
    first_major = sorted_major[0] if sorted_major else None
    papers = _paper_names(categories)
    external_certs = _external_certifications(categories)
    return {
        "code": code,
        "name": cls.get("name") or "",
        "term": cls.get("term") or "",
        "instructor": cls.get("instructor") or "",
        "rmp": cls.get("rmp") or data.get("rmp"),
        "has_reading": class_has_label(data, "Reading"),
        "has_homework": class_has_label(data, "Homework"),
        "papers": papers,
        "external_certs": external_certs,
        "has_connect_prep": class_has_connect_prep(categories),
        "has_group_project": any(c.get("is_group_project") for c in categories),
        "first_todo": first_todo,
        "first_major": first_major,
        "due_items": due_items,
    }


def _month_key(iso: str) -> str:
    return iso[:7] if len(iso) >= 7 else ""


def _month_heading(month_key: str) -> str:
    if not month_key or len(month_key) < 7:
        return "Do First This Month"
    year_s, mon_s = month_key.split("-")
    return f"Do First - {MONTH_NAMES[int(mon_s)]} {year_s}"


def group_outline_rows(
    ws,
    start_row: int,
    end_row: int,
    *,
    outline_level: int = 1,
    collapsed: bool = False,
) -> None:
    """Excel row outline so a block collapses with +/- controls (summary row stays above)."""
    if start_row > end_row:
        return
    ws.row_dimensions.group(start_row, end_row, outline_level=outline_level, hidden=collapsed)
    ws.sheet_properties.outlinePr.summaryBelow = False


def build_overview_sheet(
    wb: Workbook,
    all_data: list[dict],
    *,
    as_of: date | None = None,
    workbook_path: Path | None = None,
) -> None:
    """Cover-page stats: workload flags, first due items, monthly priority list."""
    if OVERVIEW_SHEET in wb.sheetnames:
        wb.remove(wb[OVERVIEW_SHEET])
    ws = wb.create_sheet(title=OVERVIEW_SHEET, index=0)
    ws.sheet_properties.tabColor = "203764"

    rmp_table = load_rmp_ratings(workbook_path) if workbook_path else {}

    summaries = [summarize_class(d) for d in all_data]
    summaries.sort(key=lambda s: s["code"])
    for s in summaries:
        s["rmp_entry"] = rmp_for_class(s["code"], s.get("instructor", ""), rmp_table, s.get("rmp"))

    all_due: list[dict] = []
    for s in summaries:
        all_due.extend(s["due_items"])
    all_due.sort(key=lambda x: _due_sort_key(x["due_date"], x["label"], x["name"], x["is_sub"]))

    todo_due = sorted([x for x in all_due if is_todo_item(x)], key=_todo_sort_key)
    graded_todo = [x for x in todo_due if not x.get("is_connect_prep")]
    today = as_of or OVERVIEW_AS_OF or date.today()
    focus_month = today.strftime("%Y-%m")
    do_first_graded, do_first_connect = build_do_first_items(all_due, focus_month)
    upcoming_items = upcoming_todo_items(all_due, focus_month, as_of=today)
    any_connect = any(s["has_connect_prep"] for s in summaries)

    title_fill = PatternFill("solid", fgColor="203764")
    header_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True)
    ncols = 11
    last_col = get_column_letter(ncols)

    row = 1
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value="Stats Breakdown")
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    terms = sorted({s["term"] for s in summaries if s["term"]})
    if terms:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        ws.cell(row=row, column=1, value=f"Term: {', '.join(terms)}").font = Font(italic=True)
        row += 1

    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws.cell(row=row, column=1, value=f"Today: {today.isoformat()}").font = Font(italic=True, color="404040")
    row += 1

    rmp_rows = [s for s in summaries if s.get("rmp_entry")]
    if rmp_rows:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        rmp_title = ws.cell(
            row=row,
            column=1,
            value="Professor Ratings (Rate My Professors) - click +/- at left to expand/collapse",
        )
        rmp_title.font = label_font
        rmp_title.fill = section_fill
        row += 1
        rmp_group_start = row
        rmp_headers = [
            "Class",
            "Instructor",
            "Quality (/5)",
            "Difficulty (/5)",
            "Avg Reviewer Grade",
            "Would Take Again",
            "# Ratings",
        ]
        for col, name in enumerate(rmp_headers, start=1):
            c = ws.cell(row=row, column=col, value=name)
            c.font = header_font
            c.fill = header_fill
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
        rmp_stat_rows: list[tuple[int, dict, dict]] = []
        for i, s in enumerate(rmp_rows):
            rmp = s["rmp_entry"]
            band = PatternFill("solid", fgColor="F2F2F2" if i % 2 else "FFFFFF")
            instructor = format_rmp_instructor_label(rmp, s.get("instructor", ""))
            url = rmp.get("url") or ""
            quality = rmp.get("quality")
            difficulty = rmp.get("difficulty")
            again = rmp.get("would_take_again_pct")
            num = rmp.get("num_ratings")
            avg_grade = format_avg_reviewer_grade(rmp)
            values = [
                sheet_link(sanitize_sheet_name(s["code"]), s["code"]),
                instructor,
                f"{quality:g}" if quality is not None else "-",
                f"{difficulty:g}" if difficulty is not None else "-",
                avg_grade,
                f"{again:g}%" if again is not None else "-",
                str(num) if num is not None else "-",
            ]
            for col, value in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=value)
                c.fill = band
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 1 and str(value).startswith("=HYPERLINK"):
                    c.font = Font(bold=True, color="0563C1", underline="single")
                if col == 2 and url:
                    c.hyperlink = url
                    c.font = Font(color="0563C1", underline="single")
                if col == 3 and quality is not None:
                    c.font = Font(bold=True, color="375623" if quality >= 4.0 else "C65911")
            rmp_stat_rows.append((row, s, rmp))
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:{last_col}{row}")
        detail_hdr = ws.cell(row=row, column=1, value="Consensus - recurring student review themes")
        detail_hdr.font = label_font
        detail_hdr.fill = section_fill
        row += 1
        for i, (_stat_row, s, rmp) in enumerate(rmp_stat_rows):
            band = PatternFill("solid", fgColor="F2F2F2" if i % 2 else "FFFFFF")
            code = s["code"]
            instructor = format_rmp_instructor_label(rmp, s.get("instructor", ""))
            detail = (rmp.get("consensus") or "").strip()
            ws.cell(row=row, column=1, value=code).fill = band
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=instructor).fill = band
            ws.cell(row=row, column=2).border = BORDER
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=ncols)
            detail_cell = ws.cell(row=row, column=3, value=detail)
            detail_cell.fill = band
            detail_cell.border = BORDER
            detail_cell.alignment = Alignment(wrap_text=True, vertical="top")
            line_count = max(3, detail.count("\n") + 1, (len(detail) // 90) + 1)
            ws.row_dimensions[row].height = min(180, 15 * line_count)
            row += 1

        review_classes = [
            (s, rmp)
            for (_stat_row, s, rmp) in rmp_stat_rows
            if isinstance(rmp.get("reviews"), list) and rmp["reviews"]
        ]
        if review_classes:
            row += 1
            ws.merge_cells(f"A{row}:{last_col}{row}")
            review_hdr = ws.cell(
                row=row,
                column=1,
                value="Student reviews (Rate My Professors) - newest first",
            )
            review_hdr.font = label_font
            review_hdr.fill = section_fill
            row += 1
            review_group_start = row
            review_headers = ["Class", "Date", "Grade", "Q/D", "Review"]
            for col, name in enumerate(review_headers, start=1):
                c = ws.cell(row=row, column=col, value=name)
                c.font = header_font
                c.fill = header_fill
                c.border = BORDER
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row += 1
            for i, (s, rmp) in enumerate(review_classes):
                code = s["code"]
                profile = rmp.get("rmp_instructor") or rmp.get("instructor") or ""
                reviews = sorted(
                    rmp["reviews"],
                    key=lambda r: r.get("date") or "",
                    reverse=True,
                )
                for j, rev in enumerate(reviews):
                    band = PatternFill("solid", fgColor="F2F2F2" if (i + j) % 2 else "FFFFFF")
                    q = rev.get("quality")
                    d = rev.get("difficulty")
                    qd = f"{q:g}/{d:g}" if q is not None and d is not None else "-"
                    comment = rev.get("comment") or ""
                    if rev.get("tags"):
                        tags = ", ".join(rev["tags"][:4])
                        comment = f"{comment} [{tags}]"
                    label = code if j == 0 else ""
                    instructor_note = profile if j == 0 else ""
                    ws.cell(row=row, column=1, value=label or instructor_note).fill = band
                    ws.cell(row=row, column=1).border = BORDER
                    if label:
                        ws.cell(row=row, column=1).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=rev.get("date") or "-").fill = band
                    ws.cell(row=row, column=2).border = BORDER
                    ws.cell(row=row, column=3, value=rev.get("grade") or "-").fill = band
                    ws.cell(row=row, column=3).border = BORDER
                    ws.cell(row=row, column=4, value=qd).fill = band
                    ws.cell(row=row, column=4).border = BORDER
                    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=ncols)
                    detail_cell = ws.cell(row=row, column=5, value=comment)
                    detail_cell.fill = band
                    detail_cell.border = BORDER
                    detail_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    line_count = max(2, (len(comment) // 80) + 1)
                    ws.row_dimensions[row].height = min(120, 15 * line_count)
                    row += 1
            review_group_end = row - 1
            group_outline_rows(ws, review_group_start, review_group_end, collapsed=True)
            row += 1

        rmp_as_of = next(
            (s["rmp_entry"].get("as_of") for s in rmp_rows if s["rmp_entry"].get("as_of")),
            "",
        )
        if rmp_as_of:
            ws.merge_cells(f"A{row}:{last_col}{row}")
            ws.cell(
                row=row,
                column=1,
                value=(
                    f"RMP data as of {rmp_as_of}. Source: ratemyprofessors.com. "
                    "Avg reviewer grade = mean of self-reported letter grades (4.0 scale); "
                    "not all reviewers report a grade."
                ),
            ).font = Font(italic=True, size=9, color="808080")
            row += 1
        rmp_group_end = row - 1
        group_outline_rows(ws, rmp_group_start, rmp_group_end, collapsed=True)
        row += 1

    if any_connect:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        note = ws.cell(
            row=row,
            column=1,
            value=(
                "Connect prep (e.g. BUAD-304 Personal Assessments) is ungraded class prep - "
                "not part of Participation or other graded categories."
            ),
        )
        note.font = Font(italic=True, color="0070C0")
        note.fill = PatternFill("solid", fgColor="DDEBF7")
        note.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1

    if graded_todo:
        start = graded_todo[0]
        ws.merge_cells(f"A{row}:{last_col}{row}")
        lead = (
            f"Start here (graded): {start['due_date']} - {start['class_code']} - "
            f"{start['label'] or 'Category'}: {start['name'][:70]}"
        )
        cell = ws.cell(row=row, column=1, value=lead)
        cell.font = Font(bold=True, color="C00000")
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1
    summary_headers = [
        "Class",
        "Reading",
        "Homework",
        "Papers",
        "Group Project",
        "External certs",
        "Connect prep",
        "First assignment",
        "Assign due",
        "First major",
        "Major due",
    ]
    hdr_row = row
    for col, name in enumerate(summary_headers, start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    for i, s in enumerate(summaries):
        band = PatternFill("solid", fgColor="F2F2F2" if i % 2 else "FFFFFF")
        fa = s["first_todo"]
        fm = s["first_major"]
        papers_text = ", ".join(s["papers"]) if s["papers"] else "-"
        certs_text = ", ".join(s["external_certs"]) if s["external_certs"] else "-"
        connect_text = "Ungraded prep" if s["has_connect_prep"] else "-"
        values = [
            sheet_link(sanitize_sheet_name(s["code"]), s["code"]),
            "Yes" if s["has_reading"] else "-",
            "Yes" if s["has_homework"] else "-",
            papers_text,
            "Yes" if s["has_group_project"] else "-",
            certs_text,
            connect_text,
            (fa["name"][:55] if fa else "-"),
            (fa["due_date"] if fa else "-"),
            (fm["name"][:55] if fm else "-"),
            (fm["due_date"] if fm else "-"),
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = band
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 1 and str(value).startswith("=HYPERLINK"):
                c.font = Font(bold=True, color="0563C1", underline="single")
            elif col == 6 and value != "-":
                c.font = Font(bold=True, color=SUB_LABEL_COLORS.get("Certification", "0070C0"))
            elif col == 7 and value != "-":
                c.font = Font(bold=True, italic=True, color="0070C0")
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:{last_col}{row}")
    sec = ws.cell(
        row=row,
        column=1,
        value=f"{_month_heading(focus_month)} - graded assignments (first per class always listed)",
    )
    sec.font = label_font
    sec.fill = section_fill
    row += 1

    month_headers = ["Due", "Class", "Type", "Graded?", "Item"]
    for col, name in enumerate(month_headers, start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.border = BORDER
    row += 1

    def write_todo_row(item: dict, i: int, *, connect_row: bool = False) -> None:
        nonlocal row
        band = PatternFill("solid", fgColor="F2F2F2" if i % 2 else "FFFFFF")
        label = item["label"] or ("Assignment" if item.get("is_major") else "Category")
        graded = "No" if item.get("is_connect_prep") else ("Yes" if item.get("is_graded", True) else "No")
        for col, value in enumerate(
            [item["due_date"], item["class_code"], label, graded, item["name"][:80]],
            start=1,
        ):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = band
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 3 and label in SUB_LABEL_COLORS:
                c.font = Font(bold=True, color=SUB_LABEL_COLORS[label])
            if col == 4 and value == "No":
                c.font = Font(italic=True, color="0070C0")
            if connect_row:
                c.fill = PatternFill("solid", fgColor="E8F4FC" if i % 2 else "F5FAFD")
        row += 1

    for i, item in enumerate(do_first_graded[:60]):
        write_todo_row(item, i)

    if do_first_connect:
        row += 1
        ws.merge_cells(f"A{row}:{last_col}{row}")
        conn_sec = ws.cell(
            row=row,
            column=1,
            value=f"{_month_heading(focus_month)} - Connect prep only (ungraded, not Participation)",
        )
        conn_sec.font = label_font
        conn_sec.fill = PatternFill("solid", fgColor="DDEBF7")
        row += 1
        for col, name in enumerate(month_headers, start=1):
            c = ws.cell(row=row, column=col, value=name)
            c.font = header_font
            c.fill = PatternFill("solid", fgColor="5B9BD5")
            c.border = BORDER
        row += 1
        for i, item in enumerate(do_first_connect[:40]):
            write_todo_row(item, i, connect_row=True)

    if upcoming_items:
        row += 1
        ws.merge_cells(f"A{row}:{last_col}{row}")
        up = ws.cell(row=row, column=1, value="Upcoming graded assignments (next 45 days)")
        up.font = label_font
        up.fill = section_fill
        row += 1
        for col, name in enumerate(month_headers, start=1):
            c = ws.cell(row=row, column=col, value=name)
            c.font = header_font
            c.fill = header_fill
            c.border = BORDER
        row += 1
        for i, item in enumerate(upcoming_items[:40]):
            write_todo_row(item, i)

    widths = [14, 10, 10, 22, 12, 26, 12, 14, 28, 12, 28]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["I"].width = 32
    ws.column_dimensions["C"].width = 36
    # Cover sheet scrolls freely — do not freeze RMP / summary blocks at the top.
    ws.freeze_panes = None


def refresh_overview_sheet(wb: Workbook, workbook_path: Path, *, as_of: date | None = None) -> None:
    all_data = collect_class_json_files(workbook_path)
    if len(all_data) >= 1:
        build_overview_sheet(wb, all_data, as_of=as_of, workbook_path=workbook_path)


def build(
    data: dict,
    workbook_path: Path,
    template_path: Path | None = None,
    link_base: str | None = None,
    as_of: date | None = None,
) -> None:
    cls = data.get("class", {})
    class_code = str(cls.get("code") or cls.get("name") or "Class").strip()
    sheet_name = sanitize_sheet_name(class_code)

    base_color = clean_hex(cls["color"]) if cls.get("color") else auto_color(class_code)
    header_fill = PatternFill("solid", fgColor=base_color)
    band_light = PatternFill("solid", fgColor=tint(base_color, 0.85))
    band_lighter = PatternFill("solid", fgColor=tint(base_color, 0.93))
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True)
    weight_font = Font(bold=True, size=13, color=base_color)
    weight_fill = PatternFill("solid", fgColor=tint(base_color, 0.72))

    docs_dir = workbook_path.parent / "documents"
    docs_dir.mkdir(parents=True, exist_ok=True)

    if workbook_path.exists():
        wb = load_workbook(str(workbook_path))
    else:
        tmpl = template_path or default_template()
        wb = load_workbook(str(tmpl)) if tmpl.exists() else Workbook()

    remove_old_detail_sheets(wb)
    remove_class_sheet(wb, class_code)
    remove_class_pdfs(docs_dir, class_code)

    # Generate PDF documents (full extracted syllabus info per assignment)
    pdf_paths = generate_all(data, docs_dir)
    pdf_by_slug = {p.stem: p for p in pdf_paths}

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = base_color

    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    title = ws["A1"]
    class_name = cls.get("name")
    title.value = f"{class_code}" + (f" — {class_name}" if class_name else "")
    title.font = title_font
    title.fill = header_fill
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    row = 2
    source_url = cls.get("source_url") or ""
    if source_url:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        see_cell = ws.cell(row=row, column=1)
        see_cell.value = pdf_hyperlink(source_url, "SEE HERE")
        see_cell.font = Font(bold=True, size=20, color="C00000", underline="single")
        see_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        see_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 40
        row += 1

    for label, value in [
        ("Instructor", cls.get("instructor")),
        ("Term", cls.get("term")),
    ]:
        if value:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1

    scale = data.get("grading_scale", {})
    ws.cell(row=row, column=1, value="Required for an A").font = label_font
    ws.cell(row=row, column=2, value=scale.get("a_threshold", "N/A")).font = Font(
        bold=True, color="C00000"
    )
    row += 1
    if scale.get("raw_scale"):
        ws.cell(row=row, column=1, value="Grading scale").font = label_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
        ws.cell(row=row, column=2, value=scale["raw_scale"])
        row += 1

    categories = sorted(data.get("categories", []), key=weight_sort_key)
    breakdown = weight_breakdown_text(categories)
    if breakdown:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        bar = ws.cell(row=row, column=1, value=breakdown)
        bar.font = Font(bold=True, size=11, color=base_color)
        bar.fill = PatternFill("solid", fgColor=tint(base_color, 0.78))
        bar.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1

    strategy = data.get("strategy") or cls.get("strategy")
    if strategy:
        ws.cell(row=row, column=1, value="Strategy").font = label_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
        strat_cell = ws.cell(row=row, column=2, value=strategy)
        strat_cell.alignment = Alignment(wrap_text=True, vertical="top")
        strat_cell.fill = PatternFill("solid", fgColor="E2EFDA")
        ws.row_dimensions[row].height = 48
        row += 1

    if any((c.get("name") or "") == CONNECT_PREP_MARK for c in categories):
        ws.cell(row=row, column=1, value="Connect prep").font = label_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
        conn_cell = ws.cell(
            row=row,
            column=2,
            value="Ungraded - Connect self-assessments are class prep, not part of Participation or other grades.",
        )
        conn_cell.alignment = Alignment(wrap_text=True, vertical="top")
        conn_cell.fill = PatternFill("solid", fgColor="DDEBF7")
        conn_cell.font = Font(italic=True, color="0070C0")
        ws.row_dimensions[row].height = 36
        row += 1

    row += 1

    header_row = row
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    row += 1

    for i, cat in enumerate(categories):
        slug = slugify(f"{class_code}-{cat.get('name', 'item')}")
        pdf_path = pdf_by_slug.get(slug)
        if pdf_path and link_base:
            # Hosted PDF URL — works from a standalone downloaded workbook (opens in browser).
            target = f"{link_base.rstrip('/')}/{slug}.pdf"
        elif pdf_path:
            target = f"documents/{slug}.pdf"
        else:
            target = ""

        def write_row(
            name: str,
            weight: str,
            start: str,
            due: str,
            group: str,
            details: str,
            *,
            is_sub: bool = False,
            sub_label: str = "",
        ) -> None:
            nonlocal row
            if is_sub and sub_label:
                category_value = sub_label
                details_value = name
            else:
                category_value = name
                details_value = details
            values = [category_value, weight, start, due, group, details_value]
            band = band_light if i % 2 == 0 else band_lighter
            sub_fill = PatternFill("solid", fgColor=tint(base_color, 0.96)) if is_sub else band
            for col, value in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=value)
                c.fill = sub_fill
                c.border = BORDER
                c.alignment = Alignment(
                    vertical="center",
                    wrap_text=(col in (1, ncols)),
                    horizontal="right" if col == 2 and weight and not is_sub else "left",
                )
                if col == 1 and is_sub and sub_label:
                    c.font = Font(bold=True, size=11, color=SUB_LABEL_COLORS.get(sub_label, "404040"))
                elif col == 1 and not is_sub:
                    c.font = Font(bold=True, size=11)
                if col == 2 and weight and not is_sub:
                    c.font = weight_font
                    c.fill = weight_fill
                if col == ncols and value and str(value).startswith("=HYPERLINK"):
                    c.font = Font(color="0563C1", underline="single")
            row += 1

        write_row(
            cat.get("name", ""),
            format_weight(cat),
            cat.get("start_date") or "",
            cat.get("due_date") or "",
            "Yes" if cat.get("is_group_project") else "No",
            pdf_hyperlink(target) if target else "",
        )
        for sub in sort_sub_assignments(cat.get("assignments") or []):
            sub_label = sub.get("notes") or ""
            write_row(
                sub.get("name", "Assignment"),
                sub.get("weight_note") or "",
                sub.get("start_date") or "",
                sub.get("due_date") or "",
                "",
                "",
                is_sub=True,
                sub_label=sub_label,
            )

    total = sum(float(c["weight"]) for c in categories if c.get("weight") is not None)
    if any(c.get("weight") is not None for c in categories):
        ws.cell(row=row, column=1, value="Total").font = label_font
        unit = "%" if not str(scale.get("scale_type", "")).lower().startswith("point") else " pts"
        ws.cell(row=row, column=2, value=f"{total:g}{unit}").font = weight_font

    for col, width in enumerate([22, 14, 14, 14, 14, 42], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    remove_blank_template_sheet(wb)
    refresh_overview_sheet(wb, workbook_path, as_of=as_of)
    wb.save(str(workbook_path))
    print(f"Updated sheet '{sheet_name}' in {workbook_path}")
    print(f"Wrote {len(pdf_paths)} PDF document(s) to {docs_dir}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build/append a syllabus workbook.")
    parser.add_argument("json", help="Path to the class JSON file")
    parser.add_argument("--workbook", default="syllabi.xlsx")
    parser.add_argument("--template", default=None)
    parser.add_argument(
        "--link-base",
        default=None,
        help=(
            "Base URL where PDFs are hosted (e.g. raw GitHub documents/ URL). "
            "When set, Open PDF links use full https:// URLs so a standalone "
            "downloaded workbook opens PDFs in the browser — no zip needed."
        ),
    )
    parser.add_argument(
        "--as-of-date",
        default="",
        help="Overview 'today' and focus month (YYYY-MM-DD). Defaults to system date.",
    )
    args = parser.parse_args()

    as_of: date | None = None
    if args.as_of_date:
        as_of = date.fromisoformat(args.as_of_date)

    data = json.loads(Path(args.json).expanduser().read_text(encoding="utf-8"))
    template = Path(args.template).expanduser() if args.template else None
    build(data, Path(args.workbook).expanduser(), template, args.link_base, as_of=as_of)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
